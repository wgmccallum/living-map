"""Seed loader for Coordinate Plane Schemas and KCs spreadsheet."""

from __future__ import annotations

import re
from collections import defaultdict
from pathlib import Path

import openpyxl

from .models import (
    BulkImportAnnotation,
    BulkImportData,
    BulkImportSchema,
    EdgeCreate,
    FrameCreate,
    KCCreate,
    MathConceptCreate,
)

XLSX_PATH = Path(__file__).resolve().parent.parent.parent / (
    "Coordinate Plane Schemas and KCs (relational database prototype v.3).xlsx"
)

# Schema nesting derived from Components and Features columns:
#
# COP-1000 (root)
#   COP-1200  [component of 1000]
#     COP-1210  [component of 1200]
#     COP-1220  [component of 1200]
#     COP-1230  [component of 1200]
#     COP-1400  [feature of 1200]
#   COP-1300  [feature of 1000]
#     COP-1310  [component of 1300]
#     COP-1320  [component of 1300]
#     COP-1330  [component of 1300]
#     COP-1340  [component of 1300]
#   COP-1500  [component of 1000]
#     COP-1510  [component of 1500]
#     COP-1520  [component of 1500]
# COP-2000 (root — prereqs COP-1000 at schema level)
#   COP-2100  [feature of 2000]
#   COP-2200  [feature of 2000]
#   COP-2300  [feature of 2000]

PARENT_MAP: dict[int, int | None] = {
    1000: None,
    1200: 1000,
    1210: 1200,
    1220: 1200,
    1230: 1200,
    1300: 1000,
    1310: 1300,
    1320: 1300,
    1330: 1300,
    1340: 1300,
    1400: 1200,
    1500: 1000,
    1510: 1500,
    1520: 1500,
    2000: None,
    2100: 2000,
    2200: 2000,
    2300: 2000,
}

FRAME_ID = "coordinate-plane-v1"


def _parse_language_demands(raw: str | None) -> list[str]:
    """Parse language demand string into a list of individual demands."""
    if not raw:
        return []
    # Split on comma, normalize
    demands = []
    for part in str(raw).split(","):
        part = part.strip()
        if part:
            # Capitalize first letter
            part = part[0].upper() + part[1:]
            demands.append(part)
    return demands


def _resolve_prereq(ref: str, schema_kcs: dict[float, list[dict]]) -> str | None:
    """Resolve a prereq reference like 'COPKC-1500.2' to a KC ID.

    Format: COPKC-{schema_num}.{position_1_indexed}
    """
    m = re.match(r"COPKC-(\d+)\.(\d+)", ref.strip())
    if not m:
        return None
    schema_num = float(m.group(1))
    position = int(m.group(2))
    kcs_in_schema = schema_kcs.get(schema_num, [])
    if 1 <= position <= len(kcs_in_schema):
        return kcs_in_schema[position - 1]["id"]
    return None


def load_coordinate_plane(xlsx_path: Path = XLSX_PATH) -> BulkImportData:
    """Convert the coordinate plane spreadsheet into BulkImportData."""

    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)

    # ── Read schemas ──
    ws_schemas = wb["Coordinate Plane Schema Structu"]
    schema_rows: list[dict] = []
    for row in ws_schemas.iter_rows(min_row=2, values_only=True):
        sid, cat, num, name, desc = row[0], row[1], row[2], row[3], row[4]
        notes = row[12]
        if sid and str(sid).startswith("COP"):
            schema_rows.append(
                {
                    "id": str(sid),
                    "num": int(float(num)) if num else None,
                    "name": str(name) if name else str(sid),
                    "description": str(desc) if desc else None,
                    "notes": str(notes) if notes else None,
                }
            )

    # ── Read KCs ──
    ws_kcs = wb["Coordinate Plane KCs"]
    kc_rows: list[dict] = []
    for row in ws_kcs.iter_rows(min_row=2, values_only=True):
        kid = row[0]
        if not kid or not str(kid).startswith("COP"):
            continue
        kc_rows.append(
            {
                "id": str(kid),
                "num": row[2],
                "short_desc": str(row[3]) if row[3] else str(kid),
                "long_desc": str(row[4]) if row[4] else None,
                "lang": row[5],
                "kc_type": str(row[6]) if row[6] else None,
                "prereqs": str(row[8]) if row[8] else None,
                "schema_num": float(row[10]) if row[10] else None,
                "notes": str(row[12]) if row[12] else None,
            }
        )

    # ── Group KCs by schema for prereq resolution ──
    kcs_by_schema: dict[float, list[dict]] = defaultdict(list)
    for kc in kc_rows:
        if kc["schema_num"]:
            kcs_by_schema[kc["schema_num"]].append(kc)

    # ── Build schema number → ID map ──
    schema_num_to_id = {s["num"]: s["id"] for s in schema_rows if s["num"]}

    # ── Collect all language demands ──
    all_demands: set[str] = set()
    for kc in kc_rows:
        for d in _parse_language_demands(kc["lang"]):
            all_demands.add(d)

    # ── Build KCs ──
    import_kcs: list[KCCreate] = []
    for kc in kc_rows:
        import_kcs.append(
            KCCreate(
                id=kc["id"],
                short_description=kc["short_desc"],
                long_description=kc["long_desc"],
                language_demands=_parse_language_demands(kc["lang"]),
            )
        )

    # ── Build edges from prereqs ──
    import_edges: list[EdgeCreate] = []
    for kc in kc_rows:
        if not kc["prereqs"]:
            continue
        # Could be comma-separated
        for ref in str(kc["prereqs"]).split(","):
            ref = ref.strip()
            target_id = _resolve_prereq(ref, kcs_by_schema)
            if target_id:
                import_edges.append(
                    EdgeCreate(source_kc_id=target_id, target_kc_id=kc["id"])
                )
            else:
                print(f"  WARNING: Could not resolve prereq '{ref}' for {kc['id']}")

    # ── Build frame ──
    import_frames = [
        FrameCreate(id=FRAME_ID, name="Coordinate Plane v1"),
    ]

    # ── Build schemas ──
    import_schemas: list[BulkImportSchema] = []
    for s in schema_rows:
        num = s["num"]
        parent_num = PARENT_MAP.get(num)
        parent_id = schema_num_to_id.get(parent_num) if parent_num else None

        # Find KCs belonging to this schema
        kc_ids = [kc["id"] for kc in kcs_by_schema.get(float(num), [])]

        import_schemas.append(
            BulkImportSchema(
                id=s["id"],
                frame_id=FRAME_ID,
                name=s["name"],
                description=s.get("description"),
                parent_schema_id=parent_id,
                kc_ids=kc_ids,
            )
        )

    # ── Build annotations from notes ──
    import_annotations: list[BulkImportAnnotation] = []
    for kc in kc_rows:
        if kc["notes"]:
            import_annotations.append(
                BulkImportAnnotation(
                    entity_type="kc",
                    entity_id=kc["id"],
                    annotation_type="editorial_note",
                    content=kc["notes"],
                )
            )

    # ── Build math concepts ──
    import_math_concepts = [
        MathConceptCreate(
            id="coordinate-plane",
            name="Coordinate Plane",
            description="The Euclidean plane with a coordinate system",
        ),
    ]

    # Assemble
    data = BulkImportData(
        knowledge_components=import_kcs,
        prerequisite_edges=import_edges,
        frames=import_frames,
        schemas=import_schemas,
        annotations=import_annotations,
        math_concepts=import_math_concepts,
    )

    return data


if __name__ == "__main__":
    data = load_coordinate_plane()
    print(f"KCs: {len(data.knowledge_components)}")
    print(f"Edges: {len(data.prerequisite_edges)}")
    print(f"Schemas: {len(data.schemas)}")
    print(f"Annotations: {len(data.annotations)}")
    print(f"Math concepts: {len(data.math_concepts)}")
