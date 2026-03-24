"""Seed data loader — converts the Counting Numbers spreadsheet into bulk import JSON.

Usage:
    python -m living_map.seed_loader [path_to_xlsx] [--load]

Without --load: prints the JSON to stdout.
With --load: loads directly into the database.
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl

from .models import (
    AnnotationCreate,
    BulkImportData,
    BulkImportSchema,
    EdgeCreate,
    FrameCreate,
    KCCreate,
    MathConceptCreate,
    MathContextLink,
)

DEFAULT_XLSX = Path(__file__).parent.parent.parent / "Counting Numbers Schemas and KCs Nov 2025.xlsx"

# Map spreadsheet language demand strings to normalized labels
LANGUAGE_DEMAND_MAP = {
    "speaking": "Speaking",
    "listening": "Listening",
    "reading": "Reading",
    "writing": "Writing",
    "interpreting a mathematical representation": "Interpreting a mathematical representation",
    "producing a mathematical representation": "Producing a mathematical representation",
}


def _parse_language_demands(raw: str | None) -> list[str]:
    """Parse compound language demand strings like 'Reading (receptive), Producing a mathematical representation (productive)'."""
    if not raw:
        return []
    demands = []
    # Split on comma, then strip parenthetical qualifiers
    parts = [p.strip() for p in raw.split(",")]
    for part in parts:
        # Remove parenthetical like "(productive)" or "(receptive)"
        clean = part
        if "(" in clean:
            clean = clean[: clean.index("(")].strip()
        key = clean.lower()
        if key in LANGUAGE_DEMAND_MAP:
            demands.append(LANGUAGE_DEMAND_MAP[key])
    return demands


def _parse_antecedents(raw, num_to_id: dict[int, str]) -> list[tuple[str, str]]:
    """Parse antecedent references and resolve to full KC IDs.
    Returns list of (source_kc_id, target_kc_id) pairs."""
    if not raw:
        return []
    raw_str = str(raw).replace(".0", "")
    parts = [p.strip() for p in raw_str.split(",")]
    edges = []
    for p in parts:
        try:
            num = int(float(p))
            if num in num_to_id:
                edges.append(num_to_id[num])
        except (ValueError, TypeError):
            pass
    return edges


def load_spreadsheet(xlsx_path: Path | None = None) -> BulkImportData:
    """Parse the counting numbers spreadsheet into a BulkImportData structure."""
    xlsx_path = xlsx_path or DEFAULT_XLSX
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)

    # --- Read KC sheet ---
    ws_kcs = wb["Counting Numbers KCs"]
    kc_rows = []
    num_to_id: dict[int, str] = {}
    for row in ws_kcs.iter_rows(min_row=2, values_only=True):
        kc_id = row[0]
        if not kc_id:
            continue
        kc_id = str(kc_id).strip()
        number = int(float(row[2])) if row[2] else None
        if number is not None:
            num_to_id[number] = kc_id
        kc_rows.append({
            "id": kc_id,
            "number": number,
            "short_description": str(row[3]).strip() if row[3] else "",
            "long_description": str(row[4]).strip() if row[4] else None,
            "language_demand_raw": str(row[5]).strip() if row[5] else None,
            "kc_type": str(row[6]).strip() if row[6] else None,
            "schema_number": int(float(row[7])) if row[7] else None,
            "antecedents_raw": row[8],
        })

    # --- Read Schema sheet ---
    ws_schemas = wb["Counting Numbers Schemas"]
    schema_rows = []
    schema_num_to_id: dict[int, str] = {}
    for row in ws_schemas.iter_rows(min_row=2, values_only=True):
        schema_id = row[0]
        if not schema_id:
            continue
        schema_id = str(schema_id).strip()
        number = int(float(row[2])) if row[2] else None
        parent_num = int(float(row[5])) if row[5] else None
        if number is not None:
            schema_num_to_id[number] = schema_id
        schema_rows.append({
            "id": schema_id,
            "number": number,
            "short_description": str(row[3]).strip() if row[3] else "",
            "long_description": str(row[4]).strip() if row[4] else None,
            "parent_number": parent_num,
        })

    # --- Build KCs ---
    knowledge_components = []
    annotations = []
    for kc in kc_rows:
        demands = _parse_language_demands(kc["language_demand_raw"])
        knowledge_components.append(KCCreate(
            id=kc["id"],
            short_description=kc["short_description"],
            long_description=kc["long_description"],
            language_demands=demands,
            math_contexts=[MathContextLink(math_concept_id="whole-numbers", role="primary")],
        ))
        if kc["kc_type"]:
            annotations.append(AnnotationCreate(
                entity_type="kc",
                entity_id=kc["id"],
                annotation_type="kc_type",
                content=kc["kc_type"],
                author="seed_loader",
            ))

    # --- Build prerequisite edges ---
    prerequisite_edges = []
    seen_edges = set()
    for kc in kc_rows:
        target_id = kc["id"]
        source_ids = _parse_antecedents(kc["antecedents_raw"], num_to_id)
        for source_id in source_ids:
            pair = (source_id, target_id)
            if pair not in seen_edges:
                seen_edges.add(pair)
                prerequisite_edges.append(EdgeCreate(
                    source_kc_id=source_id,
                    target_kc_id=target_id,
                ))

    # --- Build schemas ---
    frame_id = "counting-numbers-v1"
    frames = [FrameCreate(
        id=frame_id,
        name="Counting Numbers Reference Frame",
        description="Initial frame from the counting numbers seed data",
        frame_type="internal",
        is_reference=True,
    )]

    schemas = []
    # Build KC-to-schema membership
    schema_kc_map: dict[str, list[str]] = {}
    for kc in kc_rows:
        if kc["schema_number"] and kc["schema_number"] in schema_num_to_id:
            sid = schema_num_to_id[kc["schema_number"]]
            schema_kc_map.setdefault(sid, []).append(kc["id"])

    for sr in schema_rows:
        parent_id = None
        if sr["parent_number"] and sr["parent_number"] in schema_num_to_id:
            parent_id = schema_num_to_id[sr["parent_number"]]
        kc_ids = schema_kc_map.get(sr["id"], [])
        schemas.append(BulkImportSchema(
            id=sr["id"],
            frame_id=frame_id,
            name=sr["short_description"],
            description=sr["long_description"],
            parent_schema_id=parent_id,
            kc_ids=kc_ids,
        ))

    # --- Math concepts ---
    math_concepts = [MathConceptCreate(
        id="whole-numbers",
        name="Whole Numbers",
        description="The counting numbers {1, 2, 3, ...} and zero",
    )]

    wb.close()

    return BulkImportData(
        math_concepts=math_concepts,
        math_concept_edges=[],
        knowledge_components=knowledge_components,
        prerequisite_edges=prerequisite_edges,
        annotations=[
            # Convert AnnotationCreate to BulkImportAnnotation
            {
                "entity_type": a.entity_type,
                "entity_id": a.entity_id,
                "annotation_type": a.annotation_type,
                "content": a.content,
                "author": a.author,
            }
            for a in annotations
        ],
        frames=frames,
        schemas=schemas,
    )


def main():
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 and not sys.argv[1].startswith("-") else DEFAULT_XLSX
    load_flag = "--load" in sys.argv

    data = load_spreadsheet(xlsx_path)

    if load_flag:
        from .database import init_db
        from .graph_store import GraphStore
        from .dal import DAL

        conn = init_db()
        graphs = GraphStore(conn)
        dal = DAL(conn, graphs)
        result = dal.bulk_import(data)
        print(json.dumps(result, indent=2))
        conn.close()
    else:
        print(json.dumps(data.model_dump(), indent=2, default=str))


if __name__ == "__main__":
    main()
